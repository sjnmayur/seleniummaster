import configparser
import openpyxl

book = openpyxl.load_workbook("C:\\Resume\\new123 abcd.xlsx")
sheet = book["cd"]
#acsheet = book.active

config = configparser.ConfigParser()
#config.read("..\\SeleniumAutomation\\TestData\\properties.ini") this does not work
config.read("C:\\Users\\mayur\\PycharmProjects\\PythonTesting\\SeleniumAutomation\\TestData\\properties.ini")

'''
for section in config.sections():
    print(section)
    for map, value in config.items(section):
        print(map, value)
'''
'''
for section in config.sections():
    if section == "excelvalue":
        for i in range(1, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                for map, mapval in config.items(section):
                    if i==1:
                        sheet.cell(row=i, column=j).value = mapval
                        j=j+1
                break
'''
i=1
for section in config.sections():
    if section == "excelvalue":
        while i<=sheet.max_row:
            j=1
            for map, mapval in config.items(section):
                sheet.cell(row=i, column=j).value = mapval
                j=j+1
            i=i+1


for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        print(sheet.cell(row=i, column=j).value)

print("new branch sprint26 1st commit")
print("new branch sprint26 2nd commit")